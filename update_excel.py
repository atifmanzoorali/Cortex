import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('Starter-Story-Links.xlsx')
ws = wb.active

# Check if Status column exists, if not add it
if ws['B1'].value != 'Status':
    ws.insert_cols(2)  # Insert new column at B
    ws['B1'] = 'Status'
    ws['B1'].font = openpyxl.styles.Font(bold=True)

# Mark first link (row 2) as completed
ws['B2'] = 'Completed'

# Save the file
wb.save('Starter-Story-Links.xlsx')
print("Updated Excel file: Marked first link as 'Completed'")
