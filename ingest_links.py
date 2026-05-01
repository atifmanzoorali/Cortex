import json
import os
import re
from yt_dlp import YoutubeDL
from youtube_transcript_api import YouTubeTranscriptApi
import openpyxl

# Setup paths
links_file = "Starter-Story-Links.xlsx"
output_dir = "Raw_Data"

if not os.path.exists(output_dir):
    os.makedirs(output_dir)

def sanitize_filename(title, max_length=50):
    if not title:
        return ""
    invalid_chars = r'[\\/*?:"<>|]'
    sanitized = re.sub(invalid_chars, '', title)
    sanitized = sanitized.replace(' ', '_')
    return sanitized[:max_length].rstrip('_')

def get_video_id(url):
    if "v=" in url:
        return url.split("v=")[1].split("&")[0]
    elif "be/" in url:
        return url.split("be/")[1].split("?")[0]
    return None

def get_transcript(video_id):
    try:
        transcript_list = YouTubeTranscriptApi().list(video_id)
        try:
            transcript = transcript_list.find_manually_created_transcript(['en'])
        except:
            transcript = transcript_list.find_generated_transcript(['en'])
        data = transcript.fetch()
        return " ".join([item.text for item in data])
    except Exception as e:
        print(f"    Transcript error: {e}")
        return ""

# Read URLs from Excel (rows 3-6 for next 4 links)
wb = openpyxl.load_workbook(links_file)
ws = wb.active
urls_to_process = []
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if 22 <= i <= 31:  # Process rows 22-31
        if row[0] and 'youtube' in str(row[0]).lower():
            urls_to_process.append((i, str(row[0]).strip()))

print(f"Processing {len(urls_to_process)} videos (rows 3-6)...")

for row_num, url in urls_to_process:
    video_id = get_video_id(url)
    if not video_id:
        print(f"[Row {row_num}] Invalid URL: {url}")
        continue
    
    print(f"[Row {row_num}] Processing: {video_id}")
    
    try:
        # Get metadata with yt-dlp
        ydl_opts = {
            'skip_download': True,
            'quiet': True,
            'no_warnings': True
        }
        
        with YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(url, download=False)
            
            metadata = {
                'title': info.get('title'),
                'description': info.get('description'),
                'duration': info.get('duration'),
                'channel': info.get('channel'),
                'view_count': info.get('view_count'),
                'upload_date': info.get('upload_date'),
                'tags': info.get('tags', []),
                'thumbnail': info.get('thumbnail'),
                'categories': info.get('categories', [])
            }
        
        # Get transcript
        transcript_text = get_transcript(video_id)
        
        # Save as JSON with new filename format
        output = {
            'video_id': video_id,
            'metadata': metadata,
            'transcript': transcript_text
        }
        
        sanitized_title = sanitize_filename(metadata.get('title', ''))
        if sanitized_title:
            filename = f'{video_id}_{sanitized_title}.json'
        else:
            filename = f'{video_id}.json'
        
        output_path = os.path.join(output_dir, filename)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False)
        
        # Mark as completed in Excel
        ws[f'B{row_num}'] = 'Completed'
        
        print(f"    Saved: {filename}")
        
    except Exception as e:
        print(f"    Failed: {str(e)}")
        ws[f'B{row_num}'] = f'Failed: {str(e)[:50]}'

# Save Excel
wb.save(links_file)
print("\nProcessing complete. Excel file updated.")
